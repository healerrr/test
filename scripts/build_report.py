import argparse
import csv
import json
import re
from collections import defaultdict
from copy import copy
from datetime import datetime, timezone, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from shutil import copyfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


TZ = timezone(timedelta(hours=8))
SKILL_DIR = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE_XLSX = SKILL_DIR / "assets" / "default_expense_summary_template.xlsx"

DETAIL_HEADERS = [
    "序号",
    "审批单号",
    "发起日期",
    "发起人",
    "费用所属企业",
    "费用所属部门",
    "费用类型",
    "不含税金额",
    "付款合计金额",
    "费用事由",
    "审批状态",
]


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str) and value.strip().lower() == "null":
        return ""
    return value


def to_decimal(value):
    value = clean(value)
    if value == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def to_number(value):
    number = to_decimal(value)
    return int(number) if number == number.to_integral_value() else float(number)


def to_datetime(value):
    if isinstance(value, (int, float)):
        return datetime.fromtimestamp(value / 1000, tz=TZ).replace(tzinfo=None)
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                pass
    return clean(value)


def form_fields(process_instance):
    fields = {}
    for item in process_instance.get("form_component_values", []) or []:
        fields[item.get("name") or ""] = clean(item.get("value"))
    return fields


def originator_name(process_instance):
    title = clean(process_instance.get("title"))
    if isinstance(title, str):
        match = re.match(r"^(.+?)提交的", title)
        if match:
            return match.group(1)
    for record in process_instance.get("operation_records", []) or []:
        if record.get("operation_type") == "START_PROCESS_INSTANCE":
            return clean(record.get("userid"))
    return clean(process_instance.get("originator_userid"))


def approval_status(process_instance):
    status_map = {
        "RUNNING": "审批中",
        "COMPLETED": "已完成",
        "TERMINATED": "已终止",
        "CANCELED": "已撤销",
    }
    result_map = {
        "agree": "同意",
        "AGREE": "同意",
        "refuse": "拒绝",
        "REFUSE": "拒绝",
        "NONE": "",
    }
    status = clean(process_instance.get("status"))
    result = clean(process_instance.get("result"))
    status_text = status_map.get(status, status)
    result_text = result_map.get(result, result)
    return f"{status_text}/{result_text}" if result_text else status_text


def load_process_instances(details_json):
    data = json.loads(Path(details_json).read_text(encoding="utf-8"))
    instances = []
    for item in data.get("details", []):
        process_instance = item.get("process_instance", item)
        if process_instance:
            instances.append(process_instance)
    instances.sort(key=lambda item: item.get("create_time") or 0)
    return data, instances


def build_detail_rows(instances):
    rows = []
    for process_instance in instances:
        fields = form_fields(process_instance)
        rows.append(
            {
                "审批单号": clean(process_instance.get("business_id")),
                "发起日期": to_datetime(process_instance.get("create_time")),
                "发起人": originator_name(process_instance),
                "费用所属企业": clean(fields.get("费用所属企业")),
                "费用所属部门": clean(fields.get("费用所属部门")),
                "费用类型": clean(fields.get("费用类型")),
                "不含税金额": to_number(fields.get("不含税金额（普票填含税）")),
                "付款合计金额": to_number(fields.get("费用付款合计金额")),
                "费用事由": clean(fields.get("费用事由")),
                "审批状态": approval_status(process_instance),
            }
        )

    approved = [row for row in rows if row["审批状态"] == "已完成/同意"]
    for index, row in enumerate(approved, start=1):
        row["序号"] = index
    return rows, approved


def write_detail_xlsx(rows, output_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "已完成同意"
    worksheet.append(DETAIL_HEADERS)
    for row in rows:
        worksheet.append([row.get(header) for header in DETAIL_HEADERS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if DETAIL_HEADERS[cell.column - 1] == "发起日期":
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    worksheet.freeze_panes = "A2"
    if worksheet.max_row > 1:
        ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        table = Table(displayName="CompletedAgreeApprovals", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    fit_columns(worksheet)
    workbook.save(output_path)


def write_detail_csv_json(rows, csv_path, json_path):
    with Path(csv_path).open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DETAIL_HEADERS)
        writer.writeheader()
        for row in rows:
            item = row_for_text(row)
            writer.writerow(item)
    Path(json_path).write_text(
        json.dumps([row_for_text(row) for row in rows], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def row_for_text(row):
    item = dict(row)
    value = item.get("发起日期")
    if isinstance(value, datetime):
        item["发起日期"] = value.strftime("%Y-%m-%d %H:%M:%S")
    return item


def fit_columns(worksheet, max_width=60):
    for col_idx, cells in enumerate(worksheet.columns, start=1):
        width = 10
        for cell in cells:
            if cell.value is None:
                continue
            length = 19 if isinstance(cell.value, datetime) else len(str(cell.value))
            width = max(width, min(max_width, length + 2))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def copy_cell_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def ensure_template_labels(worksheet, departments, expense_types):
    total_label = "合计"
    total_col = worksheet.max_column
    total_row = worksheet.max_row

    template_types = [worksheet.cell(2, col).value for col in range(2, total_col)]
    template_depts = [worksheet.cell(row, 1).value for row in range(3, total_row)]

    for expense_type in expense_types:
        if expense_type in template_types:
            continue
        insert_at = total_col
        worksheet.insert_cols(insert_at)
        template_col = max(2, insert_at - 1)
        for row in range(1, worksheet.max_row + 1):
            copy_cell_style(worksheet.cell(row, template_col), worksheet.cell(row, insert_at))
        worksheet.cell(2, insert_at).value = expense_type
        total_col += 1
        template_types.append(expense_type)
        worksheet.cell(2, total_col).value = total_label

    for department in departments:
        if department in template_depts:
            continue
        insert_at = total_row
        template_row = max(3, insert_at - 1)
        worksheet.insert_rows(insert_at)
        for col in range(1, worksheet.max_column + 1):
            copy_cell_style(worksheet.cell(template_row, col), worksheet.cell(insert_at, col))
        worksheet.cell(insert_at, 1).value = department
        total_row += 1
        template_depts.append(department)
        worksheet.cell(total_row, 1).value = total_label

    return total_row, total_col


def build_summary(rows, metric):
    amount_field = "付款合计金额" if metric == "tax-included" else "不含税金额"
    summary = defaultdict(Decimal)
    departments = []
    expense_types = []
    seen_depts = set()
    seen_types = set()

    for row in rows:
        dept = str(row.get("费用所属部门") or "")
        expense_type = str(row.get("费用类型") or "")
        amount = to_decimal(row.get(amount_field))
        summary[(dept, expense_type)] += amount
        if dept not in seen_depts:
            seen_depts.add(dept)
            departments.append(dept)
        if expense_type not in seen_types:
            seen_types.add(expense_type)
            expense_types.append(expense_type)

    return summary, departments, expense_types


def write_template_summary(rows, template_path, output_path, year, month, metric):
    summary, departments, expense_types = build_summary(rows, metric)
    copyfile(template_path, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    worksheet.title = "费用汇总表"

    total_row, total_col = ensure_template_labels(worksheet, departments, expense_types)
    template_types = [worksheet.cell(2, col).value for col in range(2, total_col)]
    template_depts = [worksheet.cell(row, 1).value for row in range(3, total_row)]

    metric_label = "含税金额" if metric == "tax-included" else "不含税金额"
    worksheet["A1"] = f"{year}年{month}月费用审批汇总表（按部门×费用类型·{metric_label}）"

    grand_total = Decimal("0")
    col_totals = {expense_type: Decimal("0") for expense_type in template_types}

    for row_idx, dept in enumerate(template_depts, start=3):
        row_total = Decimal("0")
        for col_idx, expense_type in enumerate(template_types, start=2):
            amount = summary.get((str(dept), str(expense_type)), Decimal("0"))
            worksheet.cell(row_idx, col_idx).value = float(amount)
            worksheet.cell(row_idx, col_idx).number_format = "#,##0.00"
            row_total += amount
            col_totals[expense_type] += amount
        worksheet.cell(row_idx, total_col).value = float(row_total)
        worksheet.cell(row_idx, total_col).number_format = "#,##0.00"
        grand_total += row_total

    for col_idx, expense_type in enumerate(template_types, start=2):
        worksheet.cell(total_row, col_idx).value = float(col_totals[expense_type])
        worksheet.cell(total_row, col_idx).number_format = "#,##0.00"
    worksheet.cell(total_row, total_col).value = float(grand_total)
    worksheet.cell(total_row, total_col).number_format = "#,##0.00"

    workbook.save(output_path)
    return grand_total, len(template_depts), len(template_types)


def write_plain_summary(rows, output_path, year, month, metric):
    summary, departments, expense_types = build_summary(rows, metric)
    metric_label = "含税金额" if metric == "tax-included" else "不含税金额"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "费用汇总表"
    worksheet.append([f"{year}年{month}月费用审批汇总表（按部门×费用类型·{metric_label}）"])
    worksheet.append(["费用所属部门"] + sorted(expense_types) + ["合计"])

    grand_total = Decimal("0")
    type_totals = {expense_type: Decimal("0") for expense_type in expense_types}
    for dept in sorted(departments):
        row_total = Decimal("0")
        row = [dept]
        for expense_type in sorted(expense_types):
            amount = summary.get((dept, expense_type), Decimal("0"))
            row.append(float(amount))
            row_total += amount
            type_totals[expense_type] += amount
        row.append(float(row_total))
        worksheet.append(row)
        grand_total += row_total
    worksheet.append(["合计"] + [float(type_totals[t]) for t in sorted(expense_types)] + [float(grand_total)])

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A1"].alignment = Alignment(horizontal="center")
    for cell in worksheet[2]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in worksheet.iter_rows(min_row=3, min_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    fit_columns(worksheet)
    workbook.save(output_path)
    return grand_total, len(departments), len(expense_types)


def parse_args():
    parser = argparse.ArgumentParser(description="Build monthly DingTalk expense approval summary workbooks.")
    parser.add_argument("--details-json", required=True, help="Full DingTalk approval details JSON.")
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, required=True)
    parser.add_argument(
        "--template-xlsx",
        help="Optional workbook whose first-sheet layout/style should be reused. Defaults to the bundled skill template.",
    )
    parser.add_argument("--output-dir", default=".")
    parser.add_argument("--prefix", help="Output file prefix. Defaults to YYYY_MM.")
    parser.add_argument(
        "--metric",
        choices=["tax-included", "pre-tax"],
        default="tax-included",
        help="tax-included uses 付款合计金额; pre-tax uses 不含税金额.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    prefix = args.prefix or f"{args.year}_{args.month:02d}"

    raw_data, instances = load_process_instances(args.details_json)
    all_rows, approved_rows = build_detail_rows(instances)

    detail_xlsx = output_dir / f"{prefix}_审批明细表.xlsx"
    detail_csv = output_dir / f"{prefix}_审批明细表.csv"
    detail_json = output_dir / f"{prefix}_审批明细表.json"
    summary_xlsx = output_dir / (
        f"{prefix}_费用汇总表.xlsx"
        if args.metric == "tax-included"
        else f"{prefix}_费用汇总表_不含税.xlsx"
    )

    write_detail_xlsx(approved_rows, detail_xlsx)
    write_detail_csv_json(approved_rows, detail_csv, detail_json)

    template_xlsx = Path(args.template_xlsx) if args.template_xlsx else DEFAULT_TEMPLATE_XLSX

    if template_xlsx.exists():
        grand_total, department_count, expense_type_count = write_template_summary(
            approved_rows,
            template_xlsx,
            summary_xlsx,
            args.year,
            args.month,
            args.metric,
        )
    else:
        grand_total, department_count, expense_type_count = write_plain_summary(
            approved_rows,
            summary_xlsx,
            args.year,
            args.month,
            args.metric,
        )

    result = {
        "raw_details": len(instances),
        "all_rows": len(all_rows),
        "approved_rows": len(approved_rows),
        "department_count": department_count,
        "expense_type_count": expense_type_count,
        "grand_total": float(grand_total),
        "detail_xlsx": str(detail_xlsx.resolve()),
        "summary_xlsx": str(summary_xlsx.resolve()),
        "detail_csv": str(detail_csv.resolve()),
        "detail_json": str(detail_json.resolve()),
        "metric": args.metric,
        "template_xlsx": str(template_xlsx.resolve()) if template_xlsx.exists() else None,
    }
    print(json.dumps(result, ensure_ascii=True))


if __name__ == "__main__":
    main()
